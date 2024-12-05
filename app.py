import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
import openpyxl
import os
from io import BytesIO



# Sidebar navigation
st.sidebar.title("SRC Menu")
page = st.sidebar.radio("Go to", ["Home", "Extract DEPT Exams"])

# Department Exam extract
def exam_split():

    # Streamlit title and description
    st.title("Exam File Splitter")
    st.write("Upload an Excel file with multiple sheets, and the system will extract specific columns and save each sheet as a separate file.")

    # File uploader for the exam file
    exam_file = st.file_uploader("Upload your exam Excel file (.xlsx):", type=["xlsx"])

    if exam_file:
        try:
            # Load the Excel file
            workbook = pd.ExcelFile(exam_file)
            sheets = workbook.sheet_names  # Get all sheet names

            st.write(f"**Sheets detected:** {', '.join(sheets)}")

            # Prepare for download
            output_zip = BytesIO()  # To store files in memory
            skipped_sheets = []  # Track skipped sheets
            processed_sheets = {}  # Track processed sheets with row counts

            for sheet_name in sheets:
                # Read the sheet starting from row 15
                data = pd.read_excel(workbook, sheet_name=sheet_name, header=None, skiprows=14)

                # Check if column C (3rd column) is empty
                if data.shape[1] < 3 or data.iloc[:, 2].dropna().empty:
                    skipped_sheets.append(sheet_name)
                    continue

                # Extract specific columns by position
                extracted_data = data.iloc[:, [1, 2, 7, 8, 10, 13]]  # B, C, H, I, K, N

                # Drop rows where all selected columns are empty
                extracted_data.dropna(how='all', inplace=True)

                # Save the sheet data to an in-memory file
                output_file = BytesIO()
                with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                    extracted_data.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                # Add the in-memory file to a dictionary for download
                processed_sheets[sheet_name] = output_file.getvalue()

            # Provide results summary
            st.write("### Processing Summary")
            st.write(f"**Total number of Department(s):** {len(processed_sheets)}")
            st.write(f"**Skipped records due to error:** {len(skipped_sheets)}")
            if skipped_sheets:
                st.write(f"**Skipped Department:** {', '.join(skipped_sheets)}")

            # Provide download buttons for processed sheets
            for sheet_name, file_data in processed_sheets.items():
                st.download_button(
                    label=f"Download {sheet_name}.xlsx",
                    data=file_data,
                    file_name=f"{sheet_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        except Exception as e:
            st.error(f"An error occurred: {e}")

# Home Page: Main Application
if page == "Home":
    
    st.title("Student Result Consolidator")
    st.image("https://via.placeholder.com/800x300", caption="Your App Banner Here")
    st.write("Upload your spreadsheet below:")

    #upload file
    test_file  = st.file_uploader("Upload Test Scores File", type=["xlsx"])
    lab_file  = st.file_uploader("Upload Lab Scores File", type=["xlsx"])
    exam_file  = st.file_uploader("Upload Exam Scores File", type=["xlsx"])
    ogr_template_file = st.file_uploader("Upload OGR Template File", type=["xlsx"])

    if test_file and lab_file and  exam_file and ogr_template_file:

        #Load files into dataframes
        test_score = pd.read_excel(test_file)
        Lab_score = pd.read_excel(lab_file)
        exam_scores = pd.read_excel(exam_file)
        # Load the OGR_template using openpyxl
        ogr_template = load_workbook(ogr_template_file)
        Sheet1 = ogr_template['Sheet1']
        st.write("Files upload successfully!")


    # Merge Logic
    # if exam_file:
    if test_file and lab_file and exam_file and ogr_template_file:

    #Clean and renamme columns
        test_score.rename(columns={'Reg No': 'RegNo'}, inplace=True)
        Lab_score.rename(columns={'Reg No': 'RegNo'}, inplace=True)
        exam_scores.rename(columns={'MATRIC NO': 'RegNo'}, inplace=True)

        
        # Proceed with processing
        exam_scores['Candidates Name'] = (exam_scores['SURNAME'] + " " + exam_scores['FIRSTNAME'] + " " + exam_scores['MIDDLENAME'])
        # exam_scores[['Candidates Name', 'RegNo', 'EXAM SCORE']]

        # Merge dataframes
        merged_data = pd.merge(test_score, Lab_score, on='RegNo', how='outer')
        final_result = pd.merge(merged_data, exam_scores, on='RegNo', how='outer')

        #Reoder columns
        final_result = final_result[['Candidates Name', 'RegNo', 'TEST', 'LAB', 'EXAM SCORE']]

        # Remove rows where Registration Number is missing
        final_result = final_result[final_result['Candidates Name'].notna()]

        # Drop rows where all score columns are empty
        final_result = final_result.dropna(subset=['TEST', 'LAB', 'EXAM SCORE'], how='all')

    
        # Convert all names to uppercase and sort by Name
        final_result['Candidates Name'] = final_result['Candidates Name'].str.upper()
        final_result = final_result.sort_values(by='Candidates Name').reset_index(drop=True)



        st.write("Result Consolidation Successful! Please Preview Below")
        st.dataframe(final_result)

    # Download merged result as excel

        # Write final result into Sheet1 of OGR_template
        for row_idx, row_data in enumerate(final_result.values, start=2):  # Start writing from the 2nd row
            # for col_idx, value in enumerate(row_data, start=1):  # Write to each column
        # Extract columns from the final_result DataFrame
            name, reg_no, test, practical_score, exam_score = row_data

            # Write data to specific columns
            # Sheet1.cell(row=row_idx, column=1, value=sn)                # 1st column: SN
            Sheet1.cell(row=row_idx, column=2, value=name)              # 2nd column: Name
            Sheet1.cell(row=row_idx, column=3, value=reg_no)            # 3rd column: Registration Number
            # Leave 4th column empty
            Sheet1.cell(row=row_idx, column=5, value=test)              # 5th column: Test
            Sheet1.cell(row=row_idx, column=6, value=practical_score)   # 6th column: Practical Score (Lab) 
            Sheet1.cell(row=row_idx, column=7, value=exam_score)   # 7th column: Exam Score (Lab)  
            #Sheet1.cell(row=row_idx, column=col_idx, value=value)


        # Save the updated template into an in-memory file
        output = io.BytesIO()
        ogr_template.save(output)
        output.seek(0)
        # Request filename from user

        # Request filename from user (without file extension)
        filename = st.text_input("Enter Department Name (e.g., CSC, IFT, CYB...):")

        # Ensure a valid filename
        if filename.strip():  # Ensure filename is not empty
            complete_filename = f"{filename.strip()}.xlsx"  # Add .xlsx extension
            st.download_button(
                label="Download OGR",
                data=output,
                file_name=complete_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("You are yet to enter a valid name")
            
# Exam Split Page
elif page == "Extract DEPT Exams":
    exam_split()  # Call the function from examsplit.py





